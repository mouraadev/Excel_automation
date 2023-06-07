import openpyxl
import os
from openpyxl.styles import Font, Alignment
from datetime import *
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side


if os.path.isfile('Planilha de Produtos.xlsx'):
    book = openpyxl.load_workbook('Planilha de Produtos.xlsx')
else:
    book = openpyxl.Workbook()

if 'Produtos' in book.sheetnames:
    produto_page = book['Produtos']
else:
    produto_page = book.create_sheet('Produtos')


produto_page.title = 'Produtos'
produto_page['A1'] = 'Código'
produto_page['B1'] = 'Produto'
produto_page['C1'] = 'Preço'
produto_page['D1'] = 'Quantidade'
produto_page['E1'] = 'Horário'
produto_page['F1'] = 'Data'
             

class RegistroProdutos():
    
    def __init__(self, codigo, produto, preço, quant):
        self.codigo = codigo
        self.produto = produto
        self.preço = preço
        self.quant = quant
        
    def RegistraProduto(self):
        horario = datetime.now().strftime('%H:%M')
        data = datetime.today().strftime('%d/%m')
        informacoes = (self.codigo, self.produto, self.preço, self.quant, horario, data)
        produto_page.append(informacoes)
 
    def ExcluiProduto(self, codigo):
        for row in produto_page.iter_rows(min_row=2):
            if codigo == row[0].value:
                produto_page.delete_rows(row[0].row)
                break
            else:
                print('Código do produto não encontrado')
                break
    
    def SubstituiCelulas(self, valor_antigo, valor_novo):
        for row in produto_page.iter_rows(min_row=2):
            for cell in row:
                if cell.value == valor_antigo:
                    cell.value = valor_novo
                

# teste = RegistroProdutos(1, 'pastel', 10.0, 15)
# teste.RegistraProduto()


fontFormat = Font(name='Arial', size= 11, bold=True)
produto_page['A1'].font = fontFormat
produto_page['B1'].font = fontFormat
produto_page['C1'].font = fontFormat
produto_page['D1'].font = fontFormat
produto_page['E1'].font = fontFormat
produto_page['F1'].font = fontFormat


currency_format = '$#,##0.00'
coluna_preco = produto_page['C']
for cell in coluna_preco:
    cell.number_format = currency_format


align = Alignment(horizontal='center', vertical='center')
for row in produto_page.iter_rows():
    for cell in row:
        cell.alignment = align


cinza = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
branco = PatternFill(start_color="00FFFFFF", end_color="00FFFFFF", fill_type="solid")
contador = 0
for row in produto_page.iter_rows(min_row=2):
    for cell in row:
        if contador % 2 == 0:
            cell.fill = branco
        else:
            cell.fill = cinza
    contador += 1


cabecalho = produto_page['1']
cinza_escuro = PatternFill(start_color="00969696", end_color="00969696", fill_type="solid")
for cell in cabecalho:
    cell.fill = cinza_escuro


borda_interna = Border(
    top=Side(style='thin'),
    bottom=Side(style='thin'),
    left=Side(style='thin'),
    right=Side(style='thin')
)


for row in produto_page.iter_rows(max_row=produto_page.max_row):
    for cell in row:
        cell.border = borda_interna

book.save('Planilha de Produtos.xlsx')


